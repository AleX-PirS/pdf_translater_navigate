import fitz
import os
import re
from google_trans_new import google_translator 
import openpyxl

book = openpyxl.Workbook()
sheet = book.active
sheet.cell(row = 1, column = 1).value = 'Word'
sheet.cell(row = 1, column = 2).value = 'Part of speech'
sheet.cell(row = 1, column = 3).value = 'Transcription'
sheet.cell(row = 1, column = 4).value = 'Translate'
sheet.cell(row = 1, column = 5).value = 'Example'

translator = google_translator()  

file_name = input('Enter file name: ')

try :
  file = fitz.open(file_name + '.pdf')
except :
  print('Wrong file name.')
  quit()

page_count = file.pageCount

types = {'adj' : 'adjective', 'conj' : 'conjunction', 'phr v' : 'phrasal verb', 'phr' : 'phrase',
 'pron' : 'pronoun', 'adv' : 'adverb', 'n' : 'noun', 'pl' : 'plural', 'prep' : 'preposition', 'v' : 'verb'}

new_fh = open('buffer.txt', 'w', encoding = 'utf-8')
for i in range(page_count) :
	page_i_text = file.loadPage(i).getText("text")
	new_fh.write(page_i_text)
new_fh.close()

new_new_fh = open('buffer2.txt', 'w', encoding = 'utf-8')

count = 0
fh = open('buffer.txt', encoding = 'utf-8')
for row in fh :
	row.strip()
	if row.startswith('Photocopiable © Oxford') : continue
	if row.startswith('Name') : continue
	if row.startswith('B2  Wordlist') : continue
	if row.startswith('Here is a list of useful or new words') : continue
	if row.startswith(') all appear in the Oxford 3000') : continue
	if row.startswith('Words marked with a key') : continue
	count += 1
	if count < 12 : continue
	try:
		int(row[0])
		continue
	except: pass

	new_new_fh.write(row.rstrip())

new_new_fh.close()

fh = open('buffer2.txt', encoding = 'utf-8').read()
fh = fh.replace('?', '.')
fh = fh.replace('!', '.')
fh = fh.replace('…', '.')

words = list()
phrases = list()

rphrases = re.split('[\.]([^0-9 ])', fh)
for i in range(len(rphrases)) : 
	if i == 0 : 
		phrases.append(rphrases[i])
		continue
	elif i % 2 == 1 : continue
	phrases.append(str(rphrases[i - 1]) + str(rphrases[i]))
	if i == len(rphrases) : break

tr_count = 0
total = 0
for phrase in phrases :
	tr_count += 1
	tranc = re.findall("/(.*)/", phrase)[0]
	about = re.findall("/.*/(.+)", phrase)[0]
	if phrase.find(' phr v ') > 0 : 
		part_location = phrase.find(' phr v ')
		part = phrase[part_location + 1 : part_location + 6]
		name = phrase[:part_location - 1]	
	else : 
		part = re.findall("\s+([a-z]+)\s*/.+/", phrase)[0]
		name = re.findall("([a-z\(\) ]+)\s+[a-z]+\s*/.+/", phrase)[0].rstrip()

	word_translate = translator.translate(name,lang_tgt='ru')
	total += 1  
	if tr_count % 10 == 0 : print('10 more words have been translated, total =', total)

	word = list()
	word.insert(0, name)
	word.insert(1, part)
	word.insert(2, tranc)
	word.insert(3, word_translate)
	word.insert(4, about)	
	words.append(word)

for row in words :
	cnt = -1
	for w in list(types.keys()) :
		cnt += 1
		if w == row[1] :
			row[1] = list(types.values())[cnt]
			break

sid = 1
for word_data in words :
	sid += 1
	sheet.cell(row = sid, column = 1).value = word_data[0].capitalize()
	sheet.cell(row = sid, column = 2).value = "'-" + word_data[1].lower()
	sheet.cell(row = sid, column = 3).value = '[' + word_data[2] + ']'
	sheet.cell(row = sid, column = 4).value = word_data[3].capitalize()
	sheet.cell(row = sid, column = 5).value = word_data[4]

book.save(file_name.split('.')[0] + '.xlsx')
book.close()

os.remove('buffer.txt')
os.remove('buffer2.txt')

print('Done')